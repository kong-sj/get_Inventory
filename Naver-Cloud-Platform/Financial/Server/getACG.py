import hashlib
import hmac
import base64
import time
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.styles.alignment import Alignment

def getAccessControlGroupList(accesskey,secret_key,timestamp):
    method = "GET"# Rest API 요청 형태, 조회할때 씀
    host = "https://fin-ncloud.apigw.fin-ntruss.com"#finance
    uri = "/vserver/v2/getAccessControlGroupList?regionCode=FKR&responseFormatType=json" #? 뒤는 옵션인데 한글로 되어있고 json으로 된 것을
    url = host + uri

    secret_key = bytes(secret_key, 'UTF-8')# UTF-8로 바이트화

    message = method + " " + uri + "\n" + timestamp + "\n" + accesskey
    message = bytes(message, 'UTF-8')
    sigkey = base64.b64encode(hmac.new(secret_key, message, digestmod=hashlib.sha256).digest())
    #시그니처 키 : 시크릿키와 액세스키를 통해서 암호화시킨 값, 키값을 이용해서 -> 내가 사용할 값을 암호화시키는 것
    #암호화된 키가 왜 필요한가? -> DDos 때문, 트래픽 1조~~오면 서버 터짐. 그래서 안전장치를 검
    #액세스키와 시크릿크는 사용자만 받을 수 있고 그걸로 암호화도해서 무분별한 요청을 막는다
    headers = {#어떻게 보낼 것인지, 목적이나 주소 같은건 위에서 정했고, 이건 어떻게 보낼 것인지를 정의한다
        'Content-Type': 'application/json',
        'charset': 'UTF-8',
        'Accept': '*/*',

        'x-ncp-apigw-timestamp': timestamp,#이것들은 NCP 가이드에 있음
        'x-ncp-iam-access-key': accesskey,#https://api.ncloud-docs.com/docs/common-ncpapi
        'x-ncp-apigw-signature-v2': sigkey#요청헤더를 참조해
    }

    try:
        if method == 'GET':
            response = requests.get(url, headers=headers)#조회하는 요청을 보내겠다, heders어떻게 보낼건지,"="이건 정해진 형식
    except Exception as ex:
        print(ex)
    return response.json()#보내면 반응이 오고 이걸 json으로 봄

def getAccessControlGroupRuleList(accesskey,secret_key,timestamp,ACG_No):
    method = "GET"# Rest API 요청 형태, 조회할때 씀
    host = "https://fin-ncloud.apigw.fin-ntruss.com"#민간존
    uri = "/vserver/v2/getAccessControlGroupRuleList?regionCode=FKR&responseFormatType=json&accessControlGroupNo="+ACG_No #? 뒤는 옵션인데 한글로 되어있고 json으로 된 것을
    url = host + uri

    secret_key = bytes(secret_key, 'UTF-8')# UTF-8로 바이트화

    message = method + " " + uri + "\n" + timestamp + "\n" + accesskey
    message = bytes(message, 'UTF-8')
    sigkey = base64.b64encode(hmac.new(secret_key, message, digestmod=hashlib.sha256).digest())
    #시그니처 키 : 시크릿키와 액세스키를 통해서 암호화시킨 값, 키값을 이용해서 -> 내가 사용할 값을 암호화시키는 것
    #암호화된 키가 왜 필요한가? -> DDos 때문, 트래픽 1조~~오면 서버 터짐. 그래서 안전장치를 검
    #액세스키와 시크릿크는 사용자만 받을 수 있고 그걸로 암호화도해서 무분별한 요청을 막는다
    headers = {#어떻게 보낼 것인지, 목적이나 주소 같은건 위에서 정했고, 이건 어떻게 보낼 것인지를 정의한다
        'Content-Type': 'application/json',
        'charset': 'UTF-8',
        'Accept': '*/*',

        'x-ncp-apigw-timestamp': timestamp,#이것들은 NCP 가이드에 있음
        'x-ncp-iam-access-key': accesskey,#https://api.ncloud-docs.com/docs/common-ncpapi
        'x-ncp-apigw-signature-v2': sigkey#요청헤더를 참조해
    }

    try:
        if method == 'GET':
            response = requests.get(url, headers=headers)#조회하는 요청을 보내겠다, heders어떻게 보낼건지,"="이건 정해진 형식
    except Exception as ex:
        print(ex)
    return response.json()#보내면 반응이 오고 이걸 json으로 봄

def getAccessControlGroupDetail(accesskey,secret_key,timestamp,ACG_No):
    method = "GET"# Rest API 요청 형태, 조회할때 씀
    host = "https://fin-ncloud.apigw.fin-ntruss.com"#민간존
    uri = "/vserver/v2/getAccessControlGroupDetail?regionCode=FKR&responseFormatType=json&accessControlGroupNo="+ACG_No #? 뒤는 옵션인데 한글로 되어있고 json으로 된 것을
    url = host + uri

    secret_key = bytes(secret_key, 'UTF-8')# UTF-8로 바이트화

    message = method + " " + uri + "\n" + timestamp + "\n" + accesskey
    message = bytes(message, 'UTF-8')
    sigkey = base64.b64encode(hmac.new(secret_key, message, digestmod=hashlib.sha256).digest())
    #시그니처 키 : 시크릿키와 액세스키를 통해서 암호화시킨 값, 키값을 이용해서 -> 내가 사용할 값을 암호화시키는 것
    #암호화된 키가 왜 필요한가? -> DDos 때문, 트래픽 1조~~오면 서버 터짐. 그래서 안전장치를 검
    #액세스키와 시크릿크는 사용자만 받을 수 있고 그걸로 암호화도해서 무분별한 요청을 막는다
    headers = {#어떻게 보낼 것인지, 목적이나 주소 같은건 위에서 정했고, 이건 어떻게 보낼 것인지를 정의한다
        'Content-Type': 'application/json',
        'charset': 'UTF-8',
        'Accept': '*/*',

        'x-ncp-apigw-timestamp': timestamp,#이것들은 NCP 가이드에 있음
        'x-ncp-iam-access-key': accesskey,#https://api.ncloud-docs.com/docs/common-ncpapi
        'x-ncp-apigw-signature-v2': sigkey#요청헤더를 참조해
    }

    try:
        if method == 'GET':
            response = requests.get(url, headers=headers)#조회하는 요청을 보내겠다, heders어떻게 보낼건지,"="이건 정해진 형식
    except Exception as ex:
        print(ex)
    return response.json()#보내면 반응이 오고 이걸 json으로 봄

def getNetworkInterfaceList(accesskey,secret_key,timestamp):
    method = "GET"# Rest API 요청 형태, 조회할때 씀
    host = "https://fin-ncloud.apigw.fin-ntruss.com"#민간존
    uri = "/vserver/v2/getNetworkInterfaceList?regionCode=FKR&responseFormatType=json"#? 뒤는 옵션인데 한글로 되어있고 json으로 된 것을
    url = host + uri
    
    secret_key = bytes(secret_key, 'UTF-8')# UTF-8로 바이트화

    message = method + " " + uri + "\n" + timestamp + "\n" + accesskey
    message = bytes(message, 'UTF-8')
    sigkey = base64.b64encode(hmac.new(secret_key, message, digestmod=hashlib.sha256).digest())
    #시그니처 키 : 시크릿키와 액세스키를 통해서 암호화시킨 값, 키값을 이용해서 -> 내가 사용할 값을 암호화시키는 것
    #암호화된 키가 왜 필요한가? -> DDos 때문, 트래픽 1조~~오면 서버 터짐. 그래서 안전장치를 검
    #액세스키와 시크릿크는 사용자만 받을 수 있고 그걸로 암호화도해서 무분별한 요청을 막는다
    headers = {#어떻게 보낼 것인지, 목적이나 주소 같은건 위에서 정했고, 이건 어떻게 보낼 것인지를 정의한다
        'Content-Type': 'application/json', 
        'charset': 'UTF-8', 
        'Accept': '*/*',
        
        'x-ncp-apigw-timestamp': timestamp,#이것들은 NCP 가이드에 있음
        'x-ncp-iam-access-key': accesskey,#https://api.ncloud-docs.com/docs/common-ncpapi
        'x-ncp-apigw-signature-v2': sigkey#요청헤더를 참조해
    }

    try:
        if method == 'GET':
            response = requests.get(url, headers=headers)#조회하는 요청을 보내겠다, heders어떻게 보낼건지,"="이건 정해진 형식
    except Exception as ex:
        print(ex)
    return response.json()#보내면 반응이 오고 이걸 json으로 봄

def getServerInstanceDetail(accesskey,secret_key,timestamp,Instance_No):
    method = "GET"# Rest API 요청 형태, 조회할때 씀
    host = "https://fin-ncloud.apigw.fin-ntruss.com"#민간존
    uri = "/vserver/v2/getServerInstanceDetail?regionCode=FKR&responseFormatType=json&serverInstanceNo="+Instance_No #? 뒤는 옵션인데 한글로 되어있고 json으로 된 것을
    url = host + uri
    
    secret_key = bytes(secret_key, 'UTF-8')# UTF-8로 바이트화

    message = method + " " + uri + "\n" + timestamp + "\n" + accesskey
    message = bytes(message, 'UTF-8')
    sigkey = base64.b64encode(hmac.new(secret_key, message, digestmod=hashlib.sha256).digest())
    #시그니처 키 : 시크릿키와 액세스키를 통해서 암호화시킨 값, 키값을 이용해서 -> 내가 사용할 값을 암호화시키는 것
    #암호화된 키가 왜 필요한가? -> DDos 때문, 트래픽 1조~~오면 서버 터짐. 그래서 안전장치를 검
    #액세스키와 시크릿크는 사용자만 받을 수 있고 그걸로 암호화도해서 무분별한 요청을 막는다
    headers = {#어떻게 보낼 것인지, 목적이나 주소 같은건 위에서 정했고, 이건 어떻게 보낼 것인지를 정의한다
        'Content-Type': 'application/json', 
        'charset': 'UTF-8', 
        'Accept': '*/*',
        
        'x-ncp-apigw-timestamp': timestamp,#이것들은 NCP 가이드에 있음
        'x-ncp-iam-access-key': accesskey,#https://api.ncloud-docs.com/docs/common-ncpapi
        'x-ncp-apigw-signature-v2': sigkey#요청헤더를 참조해
    }

    try:
        if method == 'GET':
            response = requests.get(url, headers=headers)#조회하는 요청을 보내겠다, heders어떻게 보낼건지,"="이건 정해진 형식
    except Exception as ex:
        print(ex)
    return response.json()#보내면 반응이 오고 이걸 json으로 봄

def getVpcDetail(accesskey,secret_key,timestamp,Vpc_No):
    method = "GET"# Rest API 요청 형태, 조회할때 씀
    host = "https://fin-ncloud.apigw.fin-ntruss.com"#민간존
    uri = "/vpc/v2/getVpcDetail?regionCode=FKR&responseFormatType=json&vpcNo="+Vpc_No#? 뒤는 옵션인데 한글로 되어있고 json으로 된 것을
    url = host + uri
    
    secret_key = bytes(secret_key, 'UTF-8')# UTF-8로 바이트화

    message = method + " " + uri + "\n" + timestamp + "\n" + accesskey
    message = bytes(message, 'UTF-8')
    sigkey = base64.b64encode(hmac.new(secret_key, message, digestmod=hashlib.sha256).digest())
    #시그니처 키 : 시크릿키와 액세스키를 통해서 암호화시킨 값, 키값을 이용해서 -> 내가 사용할 값을 암호화시키는 것
    #암호화된 키가 왜 필요한가? -> DDos 때문, 트래픽 1조~~오면 서버 터짐. 그래서 안전장치를 검
    #액세스키와 시크릿크는 사용자만 받을 수 있고 그걸로 암호화도해서 무분별한 요청을 막는다
    headers = {#어떻게 보낼 것인지, 목적이나 주소 같은건 위에서 정했고, 이건 어떻게 보낼 것인지를 정의한다
        'Content-Type': 'application/json', 
        'charset': 'UTF-8', 
        'Accept': '*/*',
        
        'x-ncp-apigw-timestamp': timestamp,#이것들은 NCP 가이드에 있음
        'x-ncp-iam-access-key': accesskey,#https://api.ncloud-docs.com/docs/common-ncpapi
        'x-ncp-apigw-signature-v2': sigkey#요청헤더를 참조해
    }

    try:
        if method == 'GET':
            response = requests.get(url, headers=headers)#조회하는 요청을 보내겠다, heders어떻게 보낼건지,"="이건 정해진 형식
    except Exception as ex:
        print(ex)
    return response.json()#보내면 반응이 오고 이걸 json으로 봄

accesskey = "D475C0F1F549B7B991D3"
secret_key = "F4F1C65429A18377D80853EA834C9EDA5652BEB1"
timestamp = str(int(time.time() * 1000))

# ACG 리스트
Acg_All_List=getAccessControlGroupList(accesskey,secret_key,timestamp)

# NIC 리스트
Nic_List=getNetworkInterfaceList(accesskey,secret_key,timestamp)

vpc_dict = {}

# 데이터 순회
for entry in Acg_All_List['getAccessControlGroupListResponse']['accessControlGroupList']:
    vpc_no = entry['vpcNo']
    
    # VPC 번호가 딕셔너리에 없으면 새 리스트로 초기화
    if vpc_no not in vpc_dict:
        vpc_dict[vpc_no] = []
    
    # 해당 VPC에 해당하는 데이터 추가
    vpc_dict[vpc_no].append(entry)

# vpcNo별로 데이터 나누기
for Vpc_no, Acg_List in vpc_dict.items():
    # Vpn 정보 조회
    Vpc_Detail=getVpcDetail(accesskey,secret_key,timestamp,Vpc_no)
    Vpc_Name=Vpc_Detail['getVpcDetailResponse']['vpcList'][0]['vpcName']
    
    # 빈 DataFrame 생성
    empty_df = pd.DataFrame()  # 데이터가 비어 있는 DataFrame
    
    # vpcNo에 해당하는 파일 이름 설정
    file_name = f"vpc_{Vpc_Name}_access_control_groups.xlsx"
    
    # 엑셀 파일로 저장 (openpyxl 엔진 사용)
    empty_df.to_excel(file_name, index=False, engine='openpyxl')
    
    # AcgNo 별로 Acg_Rule_List 조회
    for Acg in Acg_List:
        Acg_name=Acg['accessControlGroupName']
        Acg_No=Acg['accessControlGroupNo']
        Acg_Rule_list=getAccessControlGroupRuleList(accesskey,secret_key,timestamp,Acg_No)
        
        Acg_info = []
        Instance_No_List = []
        Instance_Name_List = []
        
        for Nic in Nic_List['getNetworkInterfaceListResponse']['networkInterfaceList']:
            if Acg_No in Nic['accessControlGroupNoList']:
                if Nic['instanceNo'] not in Instance_No_List:
                    Instance_No_List.append(Nic['instanceNo'])
                    
        for Instance_No in Instance_No_List:
            Instance_Detail = getServerInstanceDetail(accesskey,secret_key,timestamp,Instance_No)
            Instance_Name_List.append(Instance_Detail['getServerInstanceDetailResponse']['serverInstanceList'][0]['serverName'])
            
        # 리스트가 비어 있으면 None 입력
        if not Instance_Name_List:
            Instance_Name_List = 'None'
        
        # Acg_Rule에 값이 존재하는지 체크 후 Rule 정리
        if Acg_Rule_list['getAccessControlGroupRuleListResponse']['totalRows'] > 0:
            for Acg_Rule in Acg_Rule_list['getAccessControlGroupRuleListResponse']['accessControlGroupRuleList']:
                
                if not Acg_Rule['ipBlock']:
                    #print(Acg_Rule['accessControlGroupSequence'])
                    Acg_Detail=getAccessControlGroupDetail(accesskey,secret_key,timestamp,Acg_Rule['accessControlGroupSequence'])
                    Sourceip = Acg_Detail['getAccessControlGroupDetailResponse']['accessControlGroupList'][0]['accessControlGroupName']
                else:
                    Sourceip = Acg_Rule['ipBlock']
                    
                Direction = Acg_Rule['accessControlGroupRuleType']['codeName']
                Protocol = Acg_Rule['protocolType']['code']
                Port = Acg_Rule['portRange']
                Comment = Acg_Rule['accessControlGroupRuleDescription']

                Acg_info.append([Acg_name, Instance_Name_List, Direction, Sourceip, Protocol, Port, Comment])
            
            df = pd.DataFrame(Acg_info, columns=['ACG Name','Instance_Name','Direction','Sourceip','Protocol','Port','Comment'])
            
            # 인덱스를 1부터 시작하도록 설정
            df.index = df.index + 1  # 기존 인덱스에 1을 더함
            
            # 새로운 시트 이름
            sheet_name = Acg_name
            
            # ExcelWriter를 사용하여 파일 열기, 기존 파일에 추가
            with pd.ExcelWriter(file_name, mode='a', engine='openpyxl') as writer:
                # DataFrame을 새로운 시트에 저장
                df.to_excel(writer, sheet_name=sheet_name, header=True)
            
            # 엑셀 파일 열기
            wb = load_workbook(file_name)
            ws = wb[sheet_name]
            
            # 열 제목 셀 범위 (첫 번째 행)
            header_cells = ws[1]
            
            # 배경색 검은색, 글자색 흰색 스타일 설정
            black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            white_font = Font(color="FFFFFF", bold=True)
            
            # 각 헤더 셀에 스타일 적용
            for cell in header_cells:
                cell.fill = black_fill
                cell.font = white_font
                
            # ACG Name 열 병합
            # ACG Name 열에서 같은 값들끼리 병합하고 가운데 정렬
            for row in range(2, ws.max_row + 1):
                acg_cell = ws.cell(row=row, column=2)  # ACG Name
                if row == 2 or ws.cell(row=row-1, column=2).value != acg_cell.value:
                    start_row = row
                if row == ws.max_row or ws.cell(row=row+1, column=2).value != acg_cell.value:
                    ws.merge_cells(start_row=start_row, start_column=2, end_row=row, end_column=2)  # 병합
                    # 가운데 정렬
                    for merged_cell in ws[start_row][1:2]:
                        merged_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Instance Name 열 병합
            # Instance Name 열에서 같은 값들끼리 병합하고 가운데 정렬
            for row in range(2, ws.max_row + 1):
                instance_cell = ws.cell(row=row, column=3)  # Instance Name
                if row == 2 or ws.cell(row=row-1, column=3).value != instance_cell.value:
                    start_row = row
                if row == ws.max_row or ws.cell(row=row+1, column=3).value != instance_cell.value:
                    ws.merge_cells(start_row=start_row, start_column=3, end_row=row, end_column=3)  # Instance Name 병합
                    # 가운데 정렬
                    for merged_cell in ws[start_row][2:3]:
                        merged_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Direction 열 병합
            # Direction 열에서 같은 값들끼리 병합하고 가운데 정렬
            for row in range(2, ws.max_row + 1):
                direction_cell = ws.cell(row=row, column=4)  # Direction
                if row == 2 or ws.cell(row=row-1, column=4).value != direction_cell.value:
                    start_row = row
                if row == ws.max_row or ws.cell(row=row+1, column=4).value != direction_cell.value:
                    ws.merge_cells(start_row=start_row, start_column=4, end_row=row, end_column=4)  # Direction 병합
                    # 가운데 정렬
                    for merged_cell in ws[start_row][3:4]:
                        merged_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 테두리 스타일 정의 (데이터 셀에 적용)
            border = Border(
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'),
                left=Side(border_style='thin'),
                right=Side(border_style='thin')
            )
            
            # 데이터 범위에 대해 테두리 추가 (헤더 제외하고 데이터에만 적용)
            for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = border
            
            # 변경된 파일 저장
            wb.save(file_name)
        
    # 엑셀 파일 열기
    wb = load_workbook(file_name)
    
    # 'Sheet1' 삭제
    del wb['Sheet1']
    
    # 변경된 파일 저장
    wb.save(file_name)